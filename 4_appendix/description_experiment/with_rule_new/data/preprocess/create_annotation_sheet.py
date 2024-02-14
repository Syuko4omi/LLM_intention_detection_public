import json
import random

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

random.seed(20221018)
# create excel sheet for annotating intention description to dev data


def load_data(path: str):
    out = []
    with open(path, "r") as f_r:
        for line in f_r:
            instance: dict = json.loads(line)
            out.append(instance)
    return out


utterances = load_data("Persuasion Face Act Prediction.jsonl")

persuasion_outcome = [
    -1 for _ in range(299)
]  # Whether persuasion was success(1), failed(0), or missing(-1) according to id
for line in utterances:
    if line["actual_donation"] == 1:
        persuasion_outcome[int(line["conversation_id"])] = 1
    else:
        persuasion_outcome[int(line["conversation_id"])] = 0

DONOR_CONV_IDS = [
    i for i, x in enumerate(persuasion_outcome) if x == 1
]  # "Donor": succeeded to persuade
NON_DONOR_CONV_IDS = [
    i for i, x in enumerate(persuasion_outcome) if x == 0
]  # "Non-donor": failed to persuade
dev_and_test_donor_conv_ids = random.sample(DONOR_CONV_IDS, 46)
train_donor_conv_ids = list(set(DONOR_CONV_IDS) - set(dev_and_test_donor_conv_ids))
dev_and_test_non_donor_conv_ids = random.sample(NON_DONOR_CONV_IDS, 13)
train_non_donor_conv_ids = list(
    set(NON_DONOR_CONV_IDS) - set(dev_and_test_non_donor_conv_ids)
)
dev_donor_conv_ids = random.sample(dev_and_test_donor_conv_ids, 23)
test_donor_conv_ids = list(set(dev_and_test_donor_conv_ids) - set(dev_donor_conv_ids))
dev_non_donor_conv_ids = random.sample(dev_and_test_non_donor_conv_ids, 6)
test_non_donor_conv_ids = list(
    set(dev_and_test_non_donor_conv_ids) - set(dev_non_donor_conv_ids)
)

train_conv_ids = train_donor_conv_ids + train_non_donor_conv_ids
dev_conv_ids = dev_donor_conv_ids + dev_non_donor_conv_ids
test_conv_ids = test_donor_conv_ids + test_non_donor_conv_ids

df = pd.read_excel("Persuasion Face Act Prediction.xlsx", engine="openpyxl")
temp = df.query(f"conversation_id in {str(dev_conv_ids)}")
temp = temp[["conversation_id", "speaker", "utterance", "true_face"]]
temp["description"] = ""
temp["appropriate"] = ""
temp["alternative_description"] = ""
temp["memo"] = ""
temp.loc[temp["true_face"] == "other", "description"] = "None"
temp.loc[temp["true_face"] == "other", "appropriate"] = "None"
temp.loc[temp["true_face"] == "other", "alternative_description"] = "None"


wb = Workbook()
ws = wb.active
ws.title = "annotation sheet"
rows = dataframe_to_rows(temp, index=False, header=True)

for row_no, row in enumerate(rows, 1):
    for col_no, value in enumerate(row, 1):
        if col_no in (
            3,
            5,
            7,
            8,
        ):
            ws.cell(row=row_no, column=col_no).alignment = Alignment(wrap_text=True)
        if row[1] == 1:
            ws.cell(row=row_no, column=col_no).fill = PatternFill(
                patternType="solid", fgColor="90ce9c"
            )
        ws.cell(row=row_no, column=col_no, value=value)
        ws.cell(row=row_no, column=col_no).border = Border(
            top=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
ws.sheet_view.zoomScale = 150


ws.column_dimensions["A"].width = 13
ws.column_dimensions["B"].width = 7
ws.column_dimensions["C"].width = 40
ws.column_dimensions["D"].width = 8
ws.column_dimensions["E"].width = 30
ws.column_dimensions["F"].width = 9
ws.column_dimensions["G"].width = 30
ws.column_dimensions["H"].width = 20

header = ws[1]
for header_cell in header:
    header_cell.fill = PatternFill(patternType="solid", fgColor="d2d2d2")
    header_cell.border = Border(
        top=Side(border_style="thin", color="000000"),
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    header_cell.font = Font(bold=True, color="000000")


wb.save("description_annotated_data.xlsx")
